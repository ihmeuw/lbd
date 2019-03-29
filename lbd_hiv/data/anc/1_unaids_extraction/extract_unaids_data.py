######################################################################################################################
## This script extracts EPP, survey, and CSAVR data from .XLXS and .XML UNAIDS files for each country (if they exist).
## It will output a separate csv by year, country, and datatype.
##
## It assumes the input directory structure is XML_BASE_DIRECTORY/year/country_dir/file, i.e. XML_BASE_DIRECTORY/2016/AFG.xlxs
## It will make a new directory for each year in OUTPUT_DIRECTORY if it doesn't exists already, i.e. OUTPUT_DIRECTORY/2016/AFG_surveillanceData.csv
###################################################################################################################################################


from bs4 import BeautifulSoup
import csv
import openpyxl
import os


XML_BASE_DIRECTORY = '<<<< FILEPATH REDACTED >>>>' # the parent directory of the year and country subdirectories
OUTPUT_DIRECTORY = '<<<< FILEPATH REDACTED >>>>' # where the extracted data should go

###################
## HELPER FUNCTIONS
###################

## Returns a BeautifulSoup object from the XML file associated with the given country code (or None if not found)
## 
## Arguments:
## - country_code: the ISO3 country code of the country should correspond with the subdirectory and XML filename)
##
## Returns:
## - a BeautifulSoup object of the country's XML file
def create_soup(filepath):
	if os.path.isfile(filepath):
		try:
			with open(filepath, encoding='utf-8') as xml_file:
				soup = BeautifulSoup(xml_file.read(), 'html.parser')
				return soup
		except Exception as e:
			print('Could not open XML file for country code' + country_code)

## Returns a void tag with the given 'property' attribute or None if not found
##
## Arguments: 
## - soup: the BeautifulSoup object
## - property: the property string to search for
##
## Return:
## - A <void> tag with the given property attribute (or None if not found)
def get_field_tag(soup, property):
	return soup.find('void', {'property' : property})

## Returns the data of an element of an array
## Arguments:
## - array_tag: the BeautifulSoup <array> tag
## - index: the index of the element in the array to extract (must be an index that exists in the array)
## - datatype: the type of data to extract (e.g. double, int, boolean, string)
##
## Return:
## - the data of the element at the given index
def extract_array_element_data(array_tag, index, datatype):
	void_tag = array_tag.findChild('void', {'index': index}, recursive=False)
	if void_tag:
		data_tag = void_tag.findChild(datatype, recursive=False)
		if data_tag:
			if datatype == 'double':
				return float(data_tag.text)
			elif datatype == 'int':
				return int(data_tag.text)
			elif datatype == 'boolean':
				return bool(data_tag.text)
			else:
				return data_tag.text

#######################
## EXTRACTION FUNCTIONS
#######################

## Extracts surveillance data (if it exists) from the given BeautifulSoup object and country code and outputs to a csv file
## The csv will be titled countryCode_surveillanceDataEPP.csv
## 
## Arguments:
## - soup: the BeautifulSoup object representing the XML file related to the given country country code
## - country_code: the ISO3 country code of the country to be extracted (should correspond with the subdirectory and XML filename)
def extract_surveillance_data(soup, country_code, year_dir):
	rows = []

	## Retrieve all the ProjectionSets
	projectionsets = soup.findChildren('', {'class': 'epp2011.core.sets.ProjectionSet'})

	## Extract data from each ProjectionSet
	for projectionset in projectionsets:
		name = get_field_tag(projectionset, 'name').findChild('string', recursive=False).text # name of the ProjectionSet

		## Years are implicitly hardcoded (1985 is the base year; year is encoded as year_index + base_year)
		start_year = 1985
		last_year = 2020
		years = [year for year in range(start_year, last_year+1)]

		sitenames_tag = get_field_tag(projectionset, 'siteNames')
		if not sitenames_tag:
			return
		sitenames_array = sitenames_tag.findChild('array', recursive=False) # array of site names
		survData_array = get_field_tag(projectionset, 'survData').findChild('array', recursive=False) # nested arrays of survey data
		survSampleSizes_array = get_field_tag(projectionset, 'survSampleSizes').findChild('array', recursive=False) # nested arrays of survey sample sizes
		siteSelected_array = get_field_tag(projectionset, 'siteSelected').findChild('array', recursive=False) # array representing whether each site was selected
		site_count = int(sitenames_array['length']) # number of sites (also how many nested arrays need to be extracted)

		for site_index in range(site_count):
			sitename = extract_array_element_data(sitenames_array, site_index, 'string') # name of the site
			sitename_survData_array = survData_array.findChild('void', {'index': site_index}, recursive=False).findChild('array', recursive=False) # array of the site's survey data
			sitename_survSampleSizes_array = survSampleSizes_array.findChild('void', {'index': site_index}, recursive=False).findChild('array', recursive=False) # array of the site's survey sample sizes
			siteSelected = extract_array_element_data(siteSelected_array, site_index, 'boolean') # whether the site was selected
			siteSelected = int(siteSelected) if siteSelected else 0 # (1 if selected, 0 otherwise) 

			for year_index in range(len(years)):
				survData = extract_array_element_data(sitename_survData_array, year_index, 'double') # extract the site's survey data for each year
				survSampleSizes = extract_array_element_data(sitename_survSampleSizes_array, year_index, 'int') # extract the site's sample size for each year

				group, site, year, prev, n, _in = name, sitename, years[year_index], survData, survSampleSizes, siteSelected # assign the csv columns
				if prev and prev != -1: # only write rows with prevalence data (-1 represents no data in the XML)
					row = [group, site, year, prev, n, _in]
					rows.append(row)
	
	## Write data to csv
	if rows:
		with open(OUTPUT_DIRECTORY + year_dir + country_code + '_surveillanceDataEPP.csv', 'w+', newline='', encoding='utf-8') as f:
			f.write('\ufeff') # write BOM to beginning of the csv file so Cyrillic characters show up correctly on Windows

			writer = csv.writer(f)
			header = ['Group', 'Site', 'Year', 'Prev', 'N', 'In']
			writer.writerow(header)
			for row in rows:
				writer.writerow(row)

## Extracts survey data (if it exists) from the given BeautifulSoup object and country code and outputs to a csv file
## The csv will be titled countryCode_surveyData.csv
## 
## Arguments:
## - soup: the BeautifulSoup object representing the XML file related to the given country country code
## - country_code: the ISO3 country code of the country to be extracted (should correspond with the subdirectory and XML filename)
def extract_survey_data(soup, country_code, year_dir):
	rows = []

	## Retrieve all the ProjectionSets
	projectionsets = soup.findChildren('', {'class': 'epp2011.core.sets.ProjectionSet'})

	## Extract each data from each ProjectionSet
	for projectionset in projectionsets:
		name = get_field_tag(projectionset, 'name').findChild('string', recursive=False).text # name of the ProjectionSet

		surveyYears_array = get_field_tag(projectionset, 'surveyYears').findChild('array') # array of when each survey was conducted
		surveyHIV_array = get_field_tag(projectionset, 'surveyHIV').findChild('array') # array of HIV percentages
		surveyStandardError_array = get_field_tag(projectionset, 'surveyStandardError').findChild('array') # array of standard error for each survey
		surveySampleSize_array = get_field_tag(projectionset, 'surveySampleSize').findChild('array') # array of survey sample sizes
		surveyIsUsed_array = get_field_tag(projectionset, 'surveyIsUsed').findChild('array') # array representing whether each survey was used

		length = int(surveyYears_array['length']) # how many elements should be in the array (some elements may be missing in the XML)

		for index in range(length): # iterate through each array concurrently
			group, year, prev, se, n, on = name, None, None, None, None, None # the csv columns

			surveyYears_child = surveyYears_array.find('void', {'index': index}) # get the year of the survey
			if surveyYears_child:
				year = surveyYears_child.findChild('int').text

			surveyHIV_child = surveyHIV_array.find('void', {'index': index}) # get prevelence
			if surveyHIV_child:
				prev = surveyHIV_child.findChild('double').text

			surveyStandardError_child = surveyStandardError_array.find('void', {'index': index}) # get the survey standard error
			if surveyStandardError_child:
				se = surveyStandardError_child.findChild('double').text

			surveySampleSize_child = surveySampleSize_array.find('void', {'index': index}) # get the survey sample size
			if surveySampleSize_child:
				n = surveySampleSize_child.findChild('int').text

			surveyIsUsed_child = surveyIsUsed_array.findChild('void', {'index': index}) # get whether the survey is used
			on = int(bool(surveyIsUsed_child.findChild('boolean').text)) if surveyIsUsed_child else 0

			if prev and prev != -1: # only write rows with prevalence data
				row = [group, year, prev, se, n, on]
				rows.append(row)
	
	## Write data to csv
	if rows:
		with open(OUTPUT_DIRECTORY + year_dir + country_code + '_surveyData.csv', 'w+', newline='', encoding='utf-8') as f:
			f.write('\ufeff') # write BOM to beginning of the csv file so Cyrillic characters show up correctly on Windows
	
			writer =  csv.writer(f)
			header = ['Group', 'Year', 'Prev', 'SE', 'N', 'On']
			writer.writerow(header)
			for row in rows:
				writer.writerow(row)

## Extracts data from the given country if the XML exists. The data will be extracted into countryCode_surveillanceData.csv and
## countryCode_surveyData.csv in the output directory set above.
##
## Arguments:
## - country_code: the ISO3 code corresponding to a subdirectory in XML_BASE_DIRECTORY (e.g. 'ZMB')
def extract_country_epp(country_code, year_dir):
	dirpath = XML_BASE_DIRECTORY + year_dir + country_code + '\\'
	if not os.path.isdir(dirpath): # make sure that dirpath actually points to a directory (not a file)
		return
	xml_files = [f for f in os.listdir(dirpath) if f.endswith('.xml')] # Find any XML files (there should really only be one)
	if not xml_files: # If the list is empty, there is no XML to extract so return
		return
	filepath = dirpath + xml_files[0] # Otherwise choose the first XML file (it should be the only XML file)
	print(filepath)

	soup = create_soup(filepath)
	if soup:
		extract_surveillance_data(soup, country_code, year_dir)
		extract_survey_data(soup, country_code, year_dir)

## Converts a worksheet row to a list of cell values
##
## Arguments:
## - worksheet: an openpyxl worksheet
## - row_index: the index of the row in the worksheet
##
## Return: returns a list of cell values
def get_row_as_list(worksheet, row_index):
	return [cell.value for cell in worksheet[row_index]]

## Extracts data from the given country if the XLXS file exists. The data will be extracted into countryCode_CSARV.csv in the output directory set above.
##
## Arguments:
## - country_code: the ISO3 code corresponding to a subdirectory in XML_BASE_DIRECTORY (e.g. 'AUS')
def extract_country_csavr(country_code, year_dir):
	dirpath = XML_BASE_DIRECTORY + year_dir + country_code + '\\'
	if not os.path.isdir(dirpath): # make sure that dirpath actually points to a directory (not a file)
		return
	xml_files = [f for f in os.listdir(dirpath) if f.endswith('.xlxs')] # Find any XLXS files (there should really only be one)
	if not xml_files: # If the list is empty, there is no XML to extract so return
		return
	input_filepath = dirpath + xml_files[0] # Otherwise choose the first XLXS file (it should be the only XLXS file)

	## This part re-writes the data to a temporary file with the correct file extension in order to not modify the input files.
	## If it's ok to modify the input files, it would probably be much faster to just rename the file to the desired extension.
	temp_filepath = OUTPUT_DIRECTORY + year_dir + country_code + '.xlsx' # a temp file with a file extension that can be read by excel (.xlSX rather than .xlXS)
	output_filepath = OUTPUT_DIRECTORY + year_dir + country_code + '_CSARV' + '.csv' # the final output csv file

	with open(input_filepath, 'rb') as input_f, open(temp_filepath, 'wb+') as output_f:
		output_f.write(input_f.read()) # writing raw input file to temp file with correct file extension (kind of a silly step, maybe there's a better way)

	worksheet = openpyxl.load_workbook(temp_filepath)['Sheet1'] # loading the temp file in as a worksheet
	row_count = worksheet.max_row + 1 # how many rows are in the file
	first_year, final_year, parent_index = None, None, None
	data_begin = 3 # the column where data starts in each row

	## extracting range of years and finding the parent tag of the data we want
	for index in range(1,row_count):
		if first_year and final_year and parent_index:
			break
		row = worksheet[index]
		for cell in row:
			if cell.value and cell.value=='<FirstYear MV>':
				first_year = int(worksheet[index+3][3].value)
				break
			elif cell.value and cell.value=='<FinalYear MV>':
				final_year = int(worksheet[index+3][3].value)
				break
			elif cell.value and cell.value=='<FitIncidenceEditorValues MV>':
				parent_index = index
				break

	if first_year and final_year and parent_index:
		## the row index of each field we want
		value_rowIndex = parent_index + 2
		plhiv_percentUndercount_rowIndex = parent_index + 3
		newHivCases_total_rowIndex = parent_index + 4
		newHivCases_undercount_rowIndex = parent_index + 5
		newHivCases_yearToDiagnosis_rowIndex = parent_index + 6
		AIDSdeaths_total_rowIndex = parent_index + 7
		AIDSdeaths_undercount_rowIndex = parent_index + 8

		## extracting each row of data
		years = [i for i in range(first_year, final_year+1)]
		value_data = get_row_as_list(worksheet, value_rowIndex)[data_begin:]
		plhiv_percentUndercount_data = get_row_as_list(worksheet, plhiv_percentUndercount_rowIndex)[data_begin:]
		newHivCases_total_data = get_row_as_list(worksheet, newHivCases_total_rowIndex)[data_begin:]
		newHivCases_undercount_data = get_row_as_list(worksheet, newHivCases_undercount_rowIndex)[data_begin:]
		newHivCases_yearToDiagnosis_data = get_row_as_list(worksheet, newHivCases_yearToDiagnosis_rowIndex)[data_begin:]
		AIDSdeaths_total_data = get_row_as_list(worksheet, AIDSdeaths_total_rowIndex)[data_begin:]
		AIDSdeaths_undercount_data = get_row_as_list(worksheet, AIDSdeaths_undercount_rowIndex)[data_begin:]

		## writing the data to csv
		with open(output_filepath, 'w+', newline='') as f:
			writer = csv.writer(f)

			## writing the header
			header = ['Year', 'PLWH_total', 'PLWH_undercount', 'new_diagnoses_total', 'new_diagnoses_years', 'new_cases_undiagnosed', 'AIDS_deaths_total', 'AIDS_deaths_undercount']
			writer.writerow(header)

			## write each year's data to csv
			for i in range(len(years)):
				## assign values to the csv columns
				year, plwh_total, plwh_undercount, new_diagnoses_total, new_diagnoses_years, new_cases_undiagnosed, AIDS_deaths_total, AIDSdeaths_undercount =\
				 years[i], value_data[i], plhiv_percentUndercount_data[i], newHivCases_total_data[i], newHivCases_yearToDiagnosis_data[i], newHivCases_undercount_data[i], \
				 AIDSdeaths_total_data[i], AIDSdeaths_undercount_data[i]

				## write the row
				writer.writerow([year, plwh_total, plwh_undercount, new_diagnoses_total, new_diagnoses_years, new_cases_undiagnosed, AIDS_deaths_total, AIDSdeaths_undercount])

		## clean up by removing the temporary file
		os.remove(temp_filepath)

## Extracts the age range the given country if the XLSX file exists. The age range will be returned.
##
## Arguments:
## - country_code: the ISO3 code corresponding to a subdirectory in XML_BASE_DIRECTORY (e.g. 'AUS')
##
## Return:
## - the age range of the sample
def extract_agerange(country_code, year_dir):
	dirpath = XML_BASE_DIRECTORY + year_dir + country_code + '\\'
	if not os.path.isdir(dirpath): # make sure that dirpath actually points to a directory (not a file)
		return
	xml_files = [f for f in os.listdir(dirpath) if f.endswith('.xlxs')] # Find any XLXS files (there should really only be one)
	if not xml_files: # If the list is empty, there is no XML to extract so return
		return
	input_filepath = dirpath + xml_files[0] # Otherwise choose the first XLXS file (it should be the only XLXS file)

	temp_filepath = OUTPUT_DIRECTORY + year_dir + country_code + '.xlsx' # a temp file with a file extension that can be read by excel (.xlSX rather than .xlXS)
	output_filepath = OUTPUT_DIRECTORY + year_dir + country_code + '_CSARV' + '.csv' # the final output csv file

	with open(input_filepath, 'rb') as input_f, open(temp_filepath, 'wb+') as output_f:
		output_f.write(input_f.read()) # writing raw input file to temp file with correct file extension (kind of a silly step, maybe there's a better way)

	workbook = openpyxl.load_workbook(temp_filepath)
	worksheet = workbook['Sheet1'] # loading the temp file in as a worksheet
	row_count = worksheet.max_row + 1 # how many rows are in the file
	
	## extracting sample age range
	row_found = False
	
	for row in worksheet:
		for index in range(worksheet.max_column):
			cell = row[index]

			## check that this row contains the header we want
			if cell.value=='Adults 15-49 = 0; Adults 15+ = 1':
				row_found = True

			## if we've found the header we want, the next non-empty cell should have the value we want
			if row_found:
				if cell.value=='0':
					age_range = 'Adults 15-49'
					os.remove(temp_filepath) # clean up by removing the temporary file
					return age_range
				elif cell.value=='1':
					age_range = 'Adults 15+'
					os.remove(temp_filepath) # clean up by removing the temporary file
					return age_range

			## if we didn't find the header, keep looping
			elif not row_found and index > 2: # optimize so that we don't have to loop through as many cells (the header is always with in the first 2 cells of the row)
				break

	## clean up by removing the temporary file
	os.remove(temp_filepath)

######################################################
######################################################

def extract_year(year_dir):
	## Make a directory for the year's extraction output if it doesn't already exist
	if not os.path.exists(OUTPUT_DIRECTORY + year_dir):
		os.makedirs(OUTPUT_DIRECTORY + year_dir)

	## Extract EPP data from all subdirectories
	for country_code in sorted(os.listdir(XML_BASE_DIRECTORY + year_dir)):
		extract_country_epp(country_code, year_dir)

	## Extracting CSAVR data
	CSAVR_countries = ['ARG', 'AUS', 'BHS', 'BLZ', 'CRI', 'CUB', 'DZA', 'GEO', 'LVA', 'MEX', 'PAN', 'SLV', 'VEN']
	for country in CSAVR_countries:
		print(country)
		extract_country_csavr(country, year_dir)

	## Extract age ranges of each country (this takes a really long time)
	with open(OUTPUT_DIRECTORY + year_dir + 'EPP_age_ranges.csv', 'a', newline='') as f:
		writer = csv.writer(f)

		header = ['Country code', 'Age range']
		writer.writerow(header)

		for country_code in sorted(os.listdir(XML_BASE_DIRECTORY + year_dir)):
			try:
				age_range = extract_agerange(country_code, year_dir)
				print(country_code + ': ' + str(age_range))
				row = [country_code, age_range if age_range else '']
				writer.writerow(row)
			except:
				pass


## Extracting code from each year's directory
for year_dir in sorted(os.listdir(XML_BASE_DIRECTORY)):
	extract_year(year_dir + '\\')
